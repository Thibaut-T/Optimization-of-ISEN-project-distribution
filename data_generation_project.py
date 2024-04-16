import random
from faker import Faker
from openpyxl import Workbook

def generate_project_data(num_rows):
    fake = Faker()
    data = []
    for i in range(1, num_rows + 1):
        numero_projet = i
        intitule = fake.word()
        propose_par = fake.first_name()
        equipe = fake.email()
        tel = None
        mail = fake.email()
        description = " ".join(fake.words(nb=10))
        min_etudiants = random.randint(2, 4)
        max_etudiants = random.randint(min_etudiants + 1, 6)
        entreprise = None
        row = [numero_projet, intitule, propose_par, equipe, tel, mail, description, min_etudiants, max_etudiants, entreprise]
        data.append(row)
    return data

def create_project_excel(filename, headers, data):
    wb = Workbook()
    ws = wb.active

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filename)
    print(f"C'est fait Iemelian '{filename}'.")

if __name__ == "__main__":
    headers = ["Project number", "Project name", "Person in charge", "Team emails", "Phone number", "Mail", 
               "Description", "Minimum students", "Maximum students", "Company"]

    data = generate_project_data(200)  
    create_project_excel("./common/dataProjects.xlsx", headers, data)
