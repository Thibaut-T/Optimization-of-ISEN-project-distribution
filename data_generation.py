import random
from faker import Faker
from openpyxl import Workbook

def generate_random_data(num_rows):
    fake = Faker()
    data = []
    for _ in range(num_rows):
        last_name = fake.last_name()
        first_name = fake.first_name()
        email = fake.email()
        state = "-"
        started_on = "-"  
        completed = "-"    
        time_taken = "-"   
        grade = "-"      
        responses = []
        for i in range(1, 63):
            response = random.randint(0, 10)
            responses.append(f"Question {i}")
            if i == 1:
                response_text = random.choice(["Non   No"]) if random.random() < 0.9 else random.choice(["Oui   Yes"])
            elif i == 2:
                response_text = random.choice(["Oui   Yes"]) if random.random() < 0.9 else random.choice(["Non   No"])
            else:
                response_text = response
            responses.append(response_text)
        row = [last_name, first_name, email, state, started_on, completed, time_taken, grade] + responses
        data.append(row)
    return data

def create_excel_file(filename, headers, data):
    wb = Workbook()
    ws = wb.active
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filename)
    print(f"Iemelian ton '{filename}' est fait.")

if __name__ == "__main__":
    headers = ["Last name", "First name", "Email address", "State", "Started on", "Completed", 
               "Time taken", "Grade/10.00"] + sum([[f"Question {i}", f"Response {i}"] for i in range(1, 63)], [])

    data = generate_random_data(200)  
    create_excel_file("./common/answerProjects.xlsx", headers, data)
